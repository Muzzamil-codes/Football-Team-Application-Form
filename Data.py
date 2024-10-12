from lib2to3.pgen2.token import NUMBER
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def FcData(list):
    if list[2] == "Team FCB":
        wfb = load_workbook("FcData.xlsx")
        wfs = wfb.active
        data = list
        for Name in wfs["A"]:
            global nar
            if Name.value == data[0]:
                print(f"Name already registered")
                nar = "True"
            else:
                nar = "False"
        for Number in wfs["B"]:
            global nur
            if Number.value == data[1]:
                print("Number already registered")
                nur = "True"
            else:
                nur = "False"
        if nur == "True" and nar == "True":
            msg = "Number and name has already been registered."
            return msg
        elif nur == "True" and nar == "False":
            msg = "Someone else has already registed with this number. Please try again with another number."
            return msg
        else:
            wfs.append(data)
            msg = "Name and number has been registered."
            wfb.save("FcData.xlsx")
            return msg
    else:
        wsb = load_workbook("SfData.xlsx")
        wss = wsb.active
        data = list
        for Name in wss["A"]:
            global narS
            if Name.value == data[0]:
                print(f"Name already registered")
                narS = "True"
            else:
                narS = "False"
        for Number in wss["B"]:
            global nurS
            if Number.value == data[1]:
                print("Number already registered")
                nurS = "True"
            else:
                nurS = "False"
        if nurS == "True" and narS == "True":
            msg = "Number and name has already been registered"
            return msg
        elif nur == "True" and nar == "False":
            msg = "Someone else has already registed with this number."
            return msg
        else:
            wss.append(data)
            msg = "Your name and number has registered"
            wsb.save("SfData.xlsx")
            return msg
